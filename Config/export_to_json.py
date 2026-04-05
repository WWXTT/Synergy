#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Excel to JSON Exporter
将 Attribute.xlsm 的数据导出为 JSON 文件

用法:
    python export_to_json.py
    python export_to_json.py --input Attribute.xlsm --output ../Assets/Configs
"""

import openpyxl
import json
import argparse
import os
import sys
from pathlib import Path


def parse_excel_to_json(xlsm_path: str) -> dict:
    """解析 Excel 文件并转换为 JSON 结构"""
    wb = openpyxl.load_workbook(xlsm_path, data_only=True)

    result = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # 查找 ##var 和 ##type 行
        var_row = None
        type_row = None
        data_start_row = 3  # 默认数据从第3行开始

        for row in range(1, min(10, ws.max_row + 1)):
            cell_a = ws.cell(row=row, column=1).value
            if cell_a == "##var":
                var_row = row
            elif cell_a == "##type":
                type_row = row

        if var_row is None or type_row is None:
            print(f"Sheet '{sheet_name}' 缺少 ##var 或 ##type 行，跳过")
            continue

        # 读取列定义
        columns = []
        for col in range(1, ws.max_column + 1):
            var_name = ws.cell(row=var_row, column=col).value
            var_type = ws.cell(row=type_row, column=col).value
            if var_name and var_name != "##var":
                columns.append({
                    "index": col,
                    "name": var_name,
                    "type": var_type or "string"
                })

        # 读取数据行
        data = []
        for row in range(data_start_row, ws.max_row + 1):
            row_data = {}
            has_data = False

            for col_info in columns:
                col = col_info["index"]
                value = ws.cell(row=row, column=col).value

                if value is not None:
                    has_data = True
                    # 类型转换
                    col_type = col_info["type"]
                    if col_type == "int" and value is not None:
                        try:
                            value = int(value)
                        except (ValueError, TypeError):
                            pass
                    elif col_type == "float" and value is not None:
                        try:
                            value = float(value)
                        except (ValueError, TypeError):
                            pass

                row_data[col_info["name"]] = value

            if has_data:
                data.append(row_data)

        result[sheet_name] = data
        print(f"Sheet '{sheet_name}': {len(data)} 行数据")

    wb.close()
    return result


def export_to_json(xlsm_path: str, output_dir: str):
    """导出 Excel 数据为 JSON 文件"""
    # 解析数据
    data = parse_excel_to_json(xlsm_path)

    # 确保输出目录存在
    output_path = Path(output_dir)
    output_path.mkdir(parents=True, exist_ok=True)

    # 为每个 sheet 生成 JSON 文件
    for sheet_name, sheet_data in data.items():
        json_file = output_path / f"{sheet_name}.json"

        with open(json_file, 'w', encoding='utf-8') as f:
            json.dump(sheet_data, f, ensure_ascii=False, indent=2)

        print(f"生成: {json_file}")

    return data


def main():
    # 默认路径
    script_dir = Path(__file__).parent.resolve()
    default_input = script_dir / "Attribute.xlsm"
    default_output = script_dir.parent / "Assets" / "Configs"

    parser = argparse.ArgumentParser(description='Excel to JSON Exporter')
    parser.add_argument('--input', '-i',
                       default=str(default_input),
                       help='输入 Excel 文件路径')
    parser.add_argument('--output', '-o',
                       default=str(default_output),
                       help='输出目录路径')

    args = parser.parse_args()

    print(f"输入文件: {args.input}")
    print(f"输出目录: {args.output}")
    print()

    if not os.path.exists(args.input):
        print(f"错误: 输入文件不存在 - {args.input}")
        sys.exit(1)

    export_to_json(args.input, args.output)
    print("\n导出完成!")


if __name__ == '__main__':
    main()